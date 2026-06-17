#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Reporte de temperatura nariz + mejillas.

Consume, por cada caso y video bajo results/<caso>/:
  - Nose information TXT/<archivo>.txt   (image, roi_temp, weighted_temp)   [nariz]
  - Cheek information TXT/<archivo>.txt   (image, L_roi_temp, L_weighted_temp,
                                           R_roi_temp, R_weighted_temp)      [mejillas]
  - Cheek coordinates TXT/<archivo>.txt   (incluye flags de calidad)

Genera 'reporte_temperatura.xlsx' con:
  - Una hoja por video: serie por frame de T_nariz, T_mej_izq, T_mej_der (C),
    diferencia nariz-mejilla, asimetria L-R y flags.
  - Hoja 'Resumen': estadisticas (media, min, max, std, n) por video.

Las temperaturas ponderadas se leen tal cual (Kelvin) y se convierten a C.

Uso:
    python generar_reporte_temperatura.py
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_RESULTS = Path(__file__).parent / "results"
OUT = Path(__file__).parent / "reporte_temperatura.xlsx"

K0 = 273.15

# ── Estilos ───────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1e3a5f")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL = PatternFill("solid", fgColor="EEF4FB")
FLAG_FILL = PatternFill("solid", fgColor="FCE4C4")
THIN = Side(style="thin", color="C0C8D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
T_FMT = "0.00"
NUM_FMT = "#,##0"


def header_row(ws, headers, widths):
    ws.row_dimensions[1].height = 28
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, ci, h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = WRAP
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w


def frame_idx(name):
    m = re.search(r"frame(\d+)", name)
    return int(m.group(1)) if m else -1


def parse_nose_info(path):
    """{frame_idx: T_nariz_C} desde Nose information (usa weighted_temp en K)."""
    out = {}
    if not path.exists():
        return out
    for line in path.read_text(encoding="utf-8").splitlines()[1:]:
        if not line.strip():
            continue
        # weighted_temp es el ultimo campo despues de la matriz ']'
        br = line.rfind("]")
        if br == -1:
            continue  # frame sin cara (solo nombre)
        name = line[: line.find(",")]
        tail = line[br + 1:].lstrip(",").strip()
        try:
            out[frame_idx(name)] = float(tail) - K0
        except ValueError:
            pass
    return out


def parse_cheek_info(path):
    """{frame_idx: (L_C, R_C)} desde Cheek information (2 weighted_temp en K)."""
    out = {}
    if not path.exists():
        return out
    for line in path.read_text(encoding="utf-8").splitlines()[1:]:
        if not line.strip():
            continue
        # estructura: name,[Lmatrix],Lwt,[Rmatrix],Rwt
        # los dos weighted estan justo despues de cada ']'
        idxs = [i for i, ch in enumerate(line) if ch == "]"]
        if len(idxs) < 2:
            continue
        name = line[: line.find(",")]
        seg_l = line[idxs[0] + 1:].lstrip(",")
        lwt = seg_l[: seg_l.find(",")]
        rwt = line[idxs[1] + 1:].lstrip(",").strip()
        try:
            out[frame_idx(name)] = (float(lwt) - K0, float(rwt) - K0)
        except ValueError:
            pass
    return out


def parse_cheek_flags(path):
    """{frame_idx: flags} desde Cheek coordinates."""
    out = {}
    if not path.exists():
        return out
    for line in path.read_text(encoding="utf-8").splitlines()[1:]:
        parts = line.split(",")
        if len(parts) < 16:
            continue
        try:
            out[int(parts[0])] = parts[-1].strip()
        except ValueError:
            pass
    return out


def stats(vals):
    vals = [v for v in vals if v is not None]
    if not vals:
        return None
    n = len(vals)
    mean = sum(vals) / n
    var = sum((v - mean) ** 2 for v in vals) / n
    return dict(n=n, mean=mean, mn=min(vals), mx=max(vals), std=var ** 0.5)


def main():
    casos = sorted(d for d in BASE_RESULTS.iterdir()
                   if d.is_dir() and (d / "Cheek information TXT").is_dir())
    if not casos:
        print(f"No hay casos con 'Cheek information TXT' en {BASE_RESULTS}")
        return

    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Resumen"
    SUM_HDR = ["Caso", "Video", "Frames",
               "Nariz med", "Nariz min", "Nariz max", "Nariz std",
               "Mej.Izq med", "Mej.Der med",
               "Nariz-Mej med", "Asim |L-R| med", "Frames marcados"]
    SUM_WID = [10, 26, 9, 11, 10, 10, 10, 12, 12, 14, 15, 16]
    header_row(ws_sum, SUM_HDR, SUM_WID)
    sum_ri = 2

    for caso in casos:
        info_dir = caso / "Cheek information TXT"
        nose_dir = caso / "Nose information TXT"
        coord_dir = caso / "Cheek coordinates TXT"

        for cheek_path in sorted(info_dir.glob("*.txt")):
            stem = cheek_path.stem
            cheeks = parse_cheek_info(cheek_path)
            noses = parse_nose_info(nose_dir / f"{stem}.txt")
            flags = parse_cheek_flags(coord_dir / f"{stem}.txt")
            if not cheeks:
                continue

            # ── Hoja por video ──
            title = f"{caso.name}_{stem}"[:31]
            ws = wb.create_sheet(title)
            HDR = ["Frame", "T_nariz", "T_mej_izq", "T_mej_der",
                   "Nariz-Mej", "Asim L-R", "Flags"]
            WID = [9, 10, 11, 11, 11, 10, 16]
            header_row(ws, HDR, WID)

            d_nm, d_asym = [], []
            ri = 2
            for idx in sorted(cheeks.keys()):
                L, R = cheeks[idx]
                nose = noses.get(idx)
                mej = (L + R) / 2.0
                nm = (nose - mej) if nose is not None else None
                asym = abs(L - R)
                d_asym.append(asym)
                if nm is not None:
                    d_nm.append(nm)
                flag = flags.get(idx, "")

                vals = [idx, nose, L, R, nm, asym, flag]
                fmts = [None, T_FMT, T_FMT, T_FMT, T_FMT, T_FMT, None]
                marked = flag and flag != "OK"
                for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
                    c = ws.cell(ri, ci, v)
                    c.border = BORDER
                    c.alignment = CENTER
                    if marked:
                        c.fill = FLAG_FILL
                    elif ri % 2 == 0:
                        c.fill = ALT_FILL
                    if fmt and v is not None:
                        c.number_format = fmt
                ri += 1
            ws.freeze_panes = "A2"

            # ── Fila de resumen ──
            sN = stats(list(noses.get(i) for i in cheeks.keys()))
            sL = stats([cheeks[i][0] for i in cheeks])
            sR = stats([cheeks[i][1] for i in cheeks])
            sNM = stats(d_nm)
            sA = stats(d_asym)
            n_marked = sum(1 for i in cheeks if flags.get(i, "OK") not in ("", "OK"))

            srow = [
                caso.name, stem, len(cheeks),
                sN["mean"] if sN else None, sN["mn"] if sN else None,
                sN["mx"] if sN else None, sN["std"] if sN else None,
                sL["mean"] if sL else None, sR["mean"] if sR else None,
                sNM["mean"] if sNM else None, sA["mean"] if sA else None,
                n_marked,
            ]
            sfmt = [None, None, NUM_FMT, T_FMT, T_FMT, T_FMT, T_FMT,
                    T_FMT, T_FMT, T_FMT, T_FMT, NUM_FMT]
            for ci, (v, fmt) in enumerate(zip(srow, sfmt), 1):
                c = ws_sum.cell(sum_ri, ci, v)
                c.border = BORDER
                c.alignment = CENTER
                if sum_ri % 2 == 0:
                    c.fill = ALT_FILL
                if fmt and v is not None:
                    c.number_format = fmt
            sum_ri += 1
            print(f"  {title}: {len(cheeks)} frames, {n_marked} marcados")

    ws_sum.freeze_panes = "A2"
    ws_sum.auto_filter.ref = f"A1:{get_column_letter(len(SUM_HDR))}1"
    wb.save(OUT)
    print(f"\nGuardado en:\n  {OUT}")


if __name__ == "__main__":
    main()
