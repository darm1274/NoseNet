#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_FN   = Path(r'D:\Usuarios\Labneurociencias\Documents\Chile\Chile\Modelos\NoseNet\data\face_no_face')
BASE_NOSE = Path(r'D:\Usuarios\Labneurociencias\Documents\Chile\Chile\Modelos\NoseNet\data\nose')
BASE_RES  = Path(r'D:\Usuarios\Labneurociencias\Documents\Chile\Chile\Modelos\NoseNet\results')
OUT       = Path(r'D:\Usuarios\Labneurociencias\Documents\Chile\Chile\Modelos\NoseNet\reporte_NoseNet.xlsx')


def parse_case(name):
    m = re.search(r'__(.*?)_(S\d+)_(P\d+)$', name)
    if m:
        return m.group(1), m.group(2), m.group(3)
    return name, '?', '?'


def confidence_stats(txt_path):
    try:
        lines = txt_path.read_text(encoding='utf-8').strip().splitlines()[1:]
        vals = [float(l.split(',')[-1].strip()) for l in lines if l.strip()]
        if vals:
            return sum(vals)/len(vals), min(vals), max(vals)
    except Exception:
        pass
    return None, None, None


# ── Recolectar datos ──────────────────────────────────────────────────────────

rows = []
cases = sorted([d.name for d in BASE_FN.iterdir()
                if d.is_dir() and d.name != 'ANT'])

for case in cases:
    grupo, persona, segmento = parse_case(case)
    face    = len(list((BASE_FN/case/'face').glob('*.png')))
    no_face = len(list((BASE_FN/case/'no_face').glob('*.png')))
    total   = face + no_face
    nose    = len(list((BASE_NOSE/case).glob('*.png')))
    conf_mean, conf_min, conf_max = confidence_stats(
        BASE_RES/case/'nose_detections.txt')

    rows.append(dict(
        grupo=grupo, persona=persona, segmento=segmento,
        total=total, face=face, no_face=no_face,
        pct_face=face/total if total else 0,
        pct_noface=no_face/total if total else 0,
        nose=nose,
        pct_nose=nose/face if face else 0,
        conf_mean=conf_mean, conf_min=conf_min, conf_max=conf_max,
    ))

print(f'Casos: {len(rows)}')

# ── Estilos comunes ───────────────────────────────────────────────────────────

HDR_FILL = PatternFill('solid', fgColor='1e3a5f')
HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
ALT_FILL = PatternFill('solid', fgColor='EEF4FB')
THIN     = Side(style='thin', color='C0C8D0')
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER   = Alignment(horizontal='center', vertical='center')
WRAP     = Alignment(horizontal='center', vertical='center', wrap_text=True)
PCT_FMT  = '0.0%'
NUM_FMT  = '#,##0'
CONF_FMT = '0.000'


def header_row(ws, headers, widths):
    ws.row_dimensions[1].height = 30
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, ci, h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = WRAP
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w


def data_cell(ws, ri, ci, val, fmt=None):
    fill = ALT_FILL if ri % 2 == 0 else None
    c = ws.cell(ri, ci, val)
    c.border = BORDER
    c.alignment = CENTER
    if fill:
        c.fill = fill
    if fmt and val is not None:
        c.number_format = fmt
    return c


# ── Hoja 1: Detalle completo ──────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = 'Detalle por caso'

COLS = [
    ('Grupo',             'grupo',      13),
    ('Persona',           'persona',    9),
    ('Segmento',          'segmento',   10),
    ('Total frames',      'total',      13),
    ('Face (n)',          'face',       10),
    ('Face (%)',          'pct_face',   10),
    ('No Face (n)',       'no_face',    12),
    ('No Face (%)',       'pct_noface', 12),
    ('Nariz det. (n)',    'nose',       14),
    ('Nariz / Face (%)', 'pct_nose',   16),
    ('Conf. media',       'conf_mean',  12),
    ('Conf. min',         'conf_min',   10),
    ('Conf. max',         'conf_max',   10),
]

header_row(ws1, [c[0] for c in COLS], [c[2] for c in COLS])

FMTS = {
    'pct_face': PCT_FMT, 'pct_noface': PCT_FMT, 'pct_nose': PCT_FMT,
    'total': NUM_FMT, 'face': NUM_FMT, 'no_face': NUM_FMT, 'nose': NUM_FMT,
    'conf_mean': CONF_FMT, 'conf_min': CONF_FMT, 'conf_max': CONF_FMT,
}

for ri, row in enumerate(rows, 2):
    for ci, (_, key, _) in enumerate(COLS, 1):
        data_cell(ws1, ri, ci, row[key], FMTS.get(key))

ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}1'


# ── Hoja 2: Resumen por Grupo ─────────────────────────────────────────────────

ws2 = wb.create_sheet('Resumen por Grupo')

HDR2 = ['Grupo', 'Casos', 'Total frames',
        'Face (n)', 'Face (%)', 'No Face (n)', 'No Face (%)',
        'Nariz det. (n)', 'Nariz / Face (%)', 'Conf. media']
WID2 = [12, 7, 13, 10, 10, 12, 12, 14, 16, 12]
header_row(ws2, HDR2, WID2)

grupos = sorted(set(r['grupo'] for r in rows))
for ri, grupo in enumerate(grupos, 2):
    g = [r for r in rows if r['grupo'] == grupo]
    tot  = sum(r['total']   for r in g)
    face = sum(r['face']    for r in g)
    nof  = sum(r['no_face'] for r in g)
    nose = sum(r['nose']    for r in g)
    cms  = [r['conf_mean']  for r in g if r['conf_mean'] is not None]
    cm   = sum(cms)/len(cms) if cms else None

    vals = [grupo, len(g), tot, face,
            face/tot if tot else 0, nof, nof/tot if tot else 0,
            nose, nose/face if face else 0, cm]
    fmts2 = [None, NUM_FMT, NUM_FMT, NUM_FMT, PCT_FMT,
             NUM_FMT, PCT_FMT, NUM_FMT, PCT_FMT, CONF_FMT]
    for ci, (val, fmt) in enumerate(zip(vals, fmts2), 1):
        data_cell(ws2, ri, ci, val, fmt)

# Fila totales
ri_tot = len(grupos) + 2
ws2.row_dimensions[ri_tot].height = 20
tot_fill = PatternFill('solid', fgColor='D0E8FF')
tot_font = Font(bold=True, size=10)
tot_tot  = sum(r['total']   for r in rows)
tot_face = sum(r['face']    for r in rows)
tot_nof  = sum(r['no_face'] for r in rows)
tot_nose = sum(r['nose']    for r in rows)
all_cms  = [r['conf_mean']  for r in rows if r['conf_mean'] is not None]
tot_cm   = sum(all_cms)/len(all_cms) if all_cms else None

totals = ['TOTAL', len(rows), tot_tot, tot_face,
          tot_face/tot_tot if tot_tot else 0, tot_nof,
          tot_nof/tot_tot if tot_tot else 0,
          tot_nose, tot_nose/tot_face if tot_face else 0, tot_cm]
tot_fmts = [None, NUM_FMT, NUM_FMT, NUM_FMT, PCT_FMT,
            NUM_FMT, PCT_FMT, NUM_FMT, PCT_FMT, CONF_FMT]
for ci, (val, fmt) in enumerate(zip(totals, tot_fmts), 1):
    c = ws2.cell(ri_tot, ci, val)
    c.fill = tot_fill
    c.font = tot_font
    c.border = BORDER
    c.alignment = CENTER
    if fmt and val is not None:
        c.number_format = fmt

ws2.freeze_panes = 'A2'


# ── Hoja 3: Resumen por Segmento ─────────────────────────────────────────────

ws3 = wb.create_sheet('Resumen por Segmento')

HDR3 = ['Segmento', 'Casos', 'Total frames',
        'Face (n)', 'Face (%)', 'No Face (n)', 'No Face (%)',
        'Nariz det. (n)', 'Nariz / Face (%)']
WID3 = [11, 7, 13, 10, 10, 12, 12, 14, 16]
header_row(ws3, HDR3, WID3)

for ri, seg in enumerate(['P1','P2','P3','P4','P5','P6'], 2):
    g = [r for r in rows if r['segmento'] == seg]
    tot  = sum(r['total']   for r in g)
    face = sum(r['face']    for r in g)
    nof  = sum(r['no_face'] for r in g)
    nose = sum(r['nose']    for r in g)
    vals = [seg, len(g), tot, face,
            face/tot if tot else 0, nof, nof/tot if tot else 0,
            nose, nose/face if face else 0]
    fmts3 = [None, NUM_FMT, NUM_FMT, NUM_FMT, PCT_FMT,
             NUM_FMT, PCT_FMT, NUM_FMT, PCT_FMT]
    for ci, (val, fmt) in enumerate(zip(vals, fmts3), 1):
        data_cell(ws3, ri, ci, val, fmt)

ws3.freeze_panes = 'A2'


# ── Hoja 4: Resumen por Persona ───────────────────────────────────────────────

ws4 = wb.create_sheet('Resumen por Persona')

HDR4 = ['Persona', 'Casos', 'Total frames',
        'Face (n)', 'Face (%)', 'No Face (n)', 'No Face (%)',
        'Nariz det. (n)', 'Nariz / Face (%)']
WID4 = [10, 7, 13, 10, 10, 12, 12, 14, 16]
header_row(ws4, HDR4, WID4)

for ri, persona in enumerate(['S1','S2'], 2):
    g = [r for r in rows if r['persona'] == persona]
    tot  = sum(r['total']   for r in g)
    face = sum(r['face']    for r in g)
    nof  = sum(r['no_face'] for r in g)
    nose = sum(r['nose']    for r in g)
    vals = [persona, len(g), tot, face,
            face/tot if tot else 0, nof, nof/tot if tot else 0,
            nose, nose/face if face else 0]
    fmts4 = [None, NUM_FMT, NUM_FMT, NUM_FMT, PCT_FMT,
             NUM_FMT, PCT_FMT, NUM_FMT, PCT_FMT]
    for ci, (val, fmt) in enumerate(zip(vals, fmts4), 1):
        data_cell(ws4, ri, ci, val, fmt)

ws4.freeze_panes = 'A2'


# ── Guardar ───────────────────────────────────────────────────────────────────

wb.save(OUT)
print(f'Guardado en:\n  {OUT}')

# ── Resumen en consola ────────────────────────────────────────────────────────

print(f'\n{"="*52}')
print(f'  RESUMEN GLOBAL — NoseNet')
print(f'{"="*52}')
print(f'  Grupos procesados  : {len(grupos)}')
print(f'  Casos totales      : {len(rows)}')
print(f'  Total frames       : {tot_tot:,}')
print(f'  Face               : {tot_face:,}  ({tot_face/tot_tot:.1%})')
print(f'  No Face            : {tot_nof:,}  ({tot_nof/tot_tot:.1%})')
print(f'  Nariz detectada    : {tot_nose:,}  ({tot_nose/tot_face:.1%} de Face)')
print(f'  Confianza media    : {tot_cm:.3f}')
print(f'{"="*52}')

# Grupos con menor deteccion de nariz
ranked = sorted(grupos, key=lambda g: sum(r['nose'] for r in rows if r['grupo']==g)
                                     / max(sum(r['face'] for r in rows if r['grupo']==g), 1))
print(f'\n  Grupos con menor deteccion de nariz:')
for g in ranked[:5]:
    gf = [r for r in rows if r['grupo'] == g]
    f  = sum(r['face'] for r in gf)
    n  = sum(r['nose'] for r in gf)
    pct = f'{n/f:.1%}' if f > 0 else 'N/A'
    print(f'    {g:<12}  {pct} ({n:,}/{f:,})')
